import random
import pandas as pd

from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

app = Flask(__name__)


def myRandom(start, stop, step):
    # Integer-based calculation avoids floating-point bias/errors
    start_i = round(start / step)
    stop_i = round(stop / step)

    return random.randint(start_i, stop_i) * step


@app.route("/", methods=["GET", "POST"])
def home():

    data = []

    # Default values
    regular_end = 66
    lateral_start = 65
    lateral_end = 75

    starting_marks = 22.0
    maximum_marks = 24.5
    total_marks = 25.0

    if request.method == "POST":

        regular_end = int(request.form["regular_end"])
        lateral_start = int(request.form["lateral_start"])
        lateral_end = int(request.form["lateral_end"])

        starting_marks = float(request.form["starting_marks"])
        maximum_marks = float(request.form["maximum_marks"])
        total_marks = float(request.form["total_marks"])


        ############### Validation #################

        # Regular roll cannot be 0
        if regular_end < 1:
            return render_template(
                "index.html",
                error="Regular roll end must be at least 1.",
                regular_end=regular_end,
                lateral_start=lateral_start,
                lateral_end=lateral_end,
                starting_marks=starting_marks,
                maximum_marks=maximum_marks,
                total_marks=total_marks
            )


        # Lateral values cannot be negative
        if lateral_start < 0 or lateral_end < 0:
            return render_template(
                "index.html",
                error="Lateral roll values cannot be negative.",
                regular_end=regular_end,
                lateral_start=lateral_start,
                lateral_end=lateral_end,
                starting_marks=starting_marks,
                maximum_marks=maximum_marks,
                total_marks=total_marks
            )


        # If lateral students exist, start cannot be greater than end
        if lateral_start > lateral_end:
            return render_template(
                "index.html",
                error="Lateral roll start cannot be greater than lateral roll end.",
                regular_end=regular_end,
                lateral_start=lateral_start,
                lateral_end=lateral_end,
                starting_marks=starting_marks,
                maximum_marks=maximum_marks,
                total_marks=total_marks
            )

        if starting_marks > maximum_marks:
            return render_template(
                "index.html",
                error="Starting marks cannot be greater than maximum marks.",
                regular_end=regular_end,
                lateral_start=lateral_start,
                lateral_end=lateral_end,
                starting_marks=starting_marks,
                maximum_marks=maximum_marks,
                total_marks=total_marks
            )

        if maximum_marks > total_marks:
            return render_template(
                "index.html",
                error="Maximum marks cannot be greater than total marks.",
                regular_end=regular_end,
                lateral_start=lateral_start,
                lateral_end=lateral_end,
                starting_marks=starting_marks,
                maximum_marks=maximum_marks,
                total_marks=total_marks
            )

        # Create roll lists
        regular_rolls = list(range(1, regular_end + 1))

        if lateral_start == 0 and lateral_end == 0:
            lateral_rolls = []
        else:
            lateral_rolls = list(range(lateral_start, lateral_end + 1))


        # Generate marks for Regular students
        for roll in regular_rolls:

            x = myRandom(
                starting_marks,
                maximum_marks,
                0.5
            )

            data.append({
                "Roll": roll,
                "Marks": round(x, 1)
            })

        # Generate marks for Lateral students
        for roll in lateral_rolls:

            x = myRandom(
                starting_marks,
                maximum_marks,
                0.5
            )

            data.append({
                "Roll": f"Lateral-{roll}",
                "Marks": round(x, 1)
            })

        # Create Excel
        df = pd.DataFrame(data)
        df.to_excel("marks.xlsx", index=False)

        # Format Excel
        wb = load_workbook("marks.xlsx")
        ws = wb.active

        thin = Side(style="thin")

        border = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        for row in ws.iter_rows():
            for cell in row:

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

                cell.border = border

        # Column width
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12

        wb.save("marks.xlsx")

    return render_template(
        "index.html",
        data=data,
        regular_end=regular_end,
        lateral_start=lateral_start,
        lateral_end=lateral_end,
        starting_marks=starting_marks,
        maximum_marks=maximum_marks,
        total_marks=total_marks
    )


@app.route("/download")
def download():

    return send_file(
        "marks.xlsx",
        as_attachment=True,
        download_name="marks.xlsx"
    )


if __name__ == "__main__":
    app.run(debug=True)
